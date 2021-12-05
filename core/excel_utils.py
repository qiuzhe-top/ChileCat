'''
Author: 邹洋
Date: 2021-07-04 13:57:48
Email: 2810201146@qq.com
LastEditors:  
LastEditTime: 2021-12-05 15:25:44
Description: Excel 操作
'''

import datetime
import os
from io import BytesIO
from typing import Pattern

import xlwt
from django.http import HttpResponse
from django.utils.encoding import escape_uri_path
from openpyxl import load_workbook

from core.settings import *


class ExcelBase:
    
    def excel_to_list(self,request):
        """Excel转列表

        Args:
            request ()): [请求]

        Returns:
            [List]]: [列表]
        """        
        file = request.data['file']
        data = []
        wb = load_workbook(file,read_only=True)
        header = None
        for rows in wb:
            for row in rows:#遍历行
                if row[0].value:
                    if not header:
                        header = self.get_header(row)
                        header_len = len(header)
                        continue
                    dict_ = {}
                    for index in range(header_len):
                        value = row[index].value
                        if value != None:
                            dict_[header[index]] = str(value)
                        else:
                            dict_[header[index]] = None
                    data.append(dict_)
        return data

    def get_header(self,row):
        '''
        excel行对象转为列表

        Args:
            row : 一行数据

        Returns:
            List: 列表
        '''
        d = []
        for k in row:
            d.append(str(k.value))
        return d

    def create_excel_response(self,name):
        '''
        创建返回

        Args:
            name (str): 文件名

        Returns:
            HttpResponse: HttpResponse对象
        '''
        response = HttpResponse(content_type='application/vnd.ms-excel')
        filename = name + datetime.date.today().strftime("%Y-%m-%d") + '.xls'
        response['Content-Disposition'] = (
            'attachment; filename={}'.format(escape_uri_path(filename))
        )
        return response
    
    def write_file(self,response,ws):
        '''
        通过IO流在网络中输出Excel文件

        Args:
            response (HttpResponse): HttpResponse对象
            ws (xlwt.Workbook): 表

        Returns:
            HttpResponse: HttpResponse对象 浏览器直接下载
        '''
        output = BytesIO()
        ws.save(output)
        output.seek(0)
        response.write(output.getvalue())
        return response

    def set_header(self,w,header):
        '''
        向w表对象第一行插入列表

        Args:
            w ([type]): [description]
            header ([type]): [description]
        '''
        column = 0
        for i in header:
            w.write(0, column, i)
            column+=1

    def download_excel(self,data,name='',header=[],type=1):
        '''
        创建excel进行下载

        Args:
            name (str): 文件名
            header (list): 列表第一行
            data (list): 数据
            type (str): 主体数据格式 1.list[list[]] 2.serializers.data[{}]
        '''
        response = self.create_excel_response(name)
        if not data:
            return
        ws,w = self.create_excel()

        #头部
        row = 0
        if len(header)>=1:
            row+=1
            self.set_header(w,header)
            
        # 主体数据
        for item in data:
            column = 0

            # 不同格式读取
            if type == 1:
                write_data = item
            elif type == 2:
                write_data = dict(item).values()

            for j in write_data:
                w.write(row, column, j)
                column += 1
            row += 1

        return self.write_file(response,ws)

    def dicts_create_excel(self,data,name='',header=[]):
        '''
        通过list创建excel进行下载

        Args:
            name (str): 文件名
            header (list): 列表第一行
            data (list): 数据
        '''
        response = self.create_excel_response(name)
        if not data:
            return
        ws,w = self.create_excel()

        #头部
        row = 0
        if len(header)>=1:
            row+=1
            self.set_header(w,header)
            
        # 主体数据
        for item in data:
            column = 0
            for j in item:
                w.write(row, column, j)
                column += 1
            row += 1

        self.write_file(response,ws)

    def create_excel(self,sheet='sheet1'):
        '''
        创建Excel对象

        Args:
            sheet (str, optional): 工作表名称. Defaults to 'sheet1'.

        Returns:
            ws,w: 下载，
        '''
        ws = xlwt.Workbook(encoding='utf-8')
        w = ws.add_sheet('sheet1')
        return ws,w
        
    def open_excel(self,path):
        '''
        打开并使用本地excel文件作为模板

        Args:
            path (str): 文件路径

        Returns:
            ws: 表sheet对象
        '''
        addr = os.getcwd()+ path
        # 打开文件
        wb = load_workbook(addr)
        # 创建一张新表
        ws = wb[wb.sheetnames[0]]
        return ws